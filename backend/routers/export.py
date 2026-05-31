import csv
import io
import logging
import zipfile
from datetime import datetime, timezone
from typing import List, Literal, Optional

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel, field_validator

from backend.core.database import get_db
from backend.core.permissions import require_admin

logger = logging.getLogger("it_admin_backend")

router = APIRouter(
    prefix="/api/v1/export",
    tags=["Export"],
    redirect_slashes=False,
)

class ExportRequest(BaseModel):
    format: Literal["csv", "xlsx"]
    types: List[Literal["assets", "users", "tasks"]]
    date_from: Optional[str] = None
    date_to: Optional[str] = None

    @field_validator("types")
    @classmethod
    def types_not_empty(cls, v):
        if not v:
            raise ValueError("Выберите хотя бы один тип данных")
        return v

ASSET_COLUMNS = [
    ("name",              "Название"),
    ("inventory_number",  "Инв. номер"),
    ("asset_type",        "Тип"),
    ("serial_number",     "Серийный номер"),
    ("mol_name",          "МОЛ"),
    ("mac_address",       "MAC-адрес"),
    ("location",          "Местоположение"),
    ("status",            "Статус"),
    ("commission_date",   "Дата ввода"),
    ("warranty_end_date", "Гарантия до"),
    ("warranty_months",   "Гарантия (мес.)"),
    ("comments",          "Комментарии"),
    ("created_at",        "Дата создания"),
]

USER_COLUMNS = [
    ("username",  "Логин"),
    ("full_name", "ФИО"),
    ("email",     "Email"),
    ("phone",     "Телефон"),
    ("role",      "Роль"),
    ("is_active", "Активен"),
]

TASK_COLUMNS = [
    ("title",            "Название"),
    ("description",      "Описание"),
    ("status",           "Статус"),
    ("priority",         "Приоритет"),
    ("task_type",        "Тип задачи"),
    ("assigned_to_name", "Исполнитель"),
    ("start_date",       "Дата начала"),
    ("due_date",         "Срок"),
    ("created_at",       "Дата создания"),
]

COLLECTION_MAP = {
    "assets": ("assets", ASSET_COLUMNS),
    "users":  ("users",  USER_COLUMNS),
    "tasks":  ("tasks",  TASK_COLUMNS),
}

def _serialize(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)

def _build_date_filter(date_from: Optional[str], date_to: Optional[str]) -> dict:
    f = {}
    if date_from:
        dt = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        f.setdefault("created_at", {})["$gte"] = dt
    if date_to:
        dt = datetime.strptime(date_to, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, tzinfo=timezone.utc
        )
        f.setdefault("created_at", {})["$lte"] = dt
    return f

async def _fetch_rows(db, collection: str, columns: list, date_filter: dict):
    headers = [col[1] for col in columns]
    fields = [col[0] for col in columns]
    docs = await db[collection].find(date_filter).to_list(length=None)
    rows = []
    for doc in docs:
        rows.append([_serialize(doc.get(f)) for f in fields])
    return headers, rows

def _build_xlsx(sheets: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append(row)

    wb.active = wb.worksheets[0]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _build_csv_single(headers: list, rows: list) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")

def _build_csv_zip(sheets: dict) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, (headers, rows) in sheets.items():
            csv_bytes = _build_csv_single(headers, rows)
            zf.writestr(f"{filename}.csv", csv_bytes)
    return buf.getvalue()

@router.post("")
async def export_data(
    req: ExportRequest,
    current_user: dict = Depends(require_admin),
    db: AsyncIOMotorDatabase = Depends(get_db),
):
    date_filter = _build_date_filter(req.date_from, req.date_to)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    sheets = {}
    sheet_labels = {"assets": "Активы", "users": "Пользователи", "tasks": "Задачи"}
    for t in req.types:
        collection, columns = COLLECTION_MAP[t]
        headers, rows = await _fetch_rows(db, collection, columns, date_filter)
        sheets[sheet_labels[t]] = (headers, rows)

    if req.format == "xlsx":
        content = _build_xlsx(sheets)
        filename = f"export_{timestamp}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    else:
        if len(sheets) == 1:
            label = next(iter(sheets))
            headers, rows = sheets[label]
            content = _build_csv_single(headers, rows)
            filename = f"export_{timestamp}.csv"
            media_type = "text/csv; charset=utf-8-sig"
        else:

            type_names = {"Активы": "assets", "Пользователи": "users", "Задачи": "tasks"}
            zip_sheets = {type_names[k]: v for k, v in sheets.items()}
            content = _build_csv_zip(zip_sheets)
            filename = f"export_{timestamp}.zip"
            media_type = "application/zip"

    logger.info("Экспорт: user=%s format=%s types=%s", current_user["id"], req.format, req.types)

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
